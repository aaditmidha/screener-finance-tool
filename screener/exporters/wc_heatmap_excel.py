"""Excel export of the working-capital heatmap with colour scales.

Writes the DSO/DIO/DPO/CCC-by-period matrix produced by
:func:`screener.models.working_capital.heatmap_data` into a worksheet and
applies a green→red colour scale per metric row (more days locked up = worse =
redder), so operational deterioration is visible at a glance in Excel itself.
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from screener.config import CONFIG

logger = logging.getLogger(__name__)

_cfg = CONFIG["exporters"]["excel"]

# Row order and display labels for the heatmap matrix.
_METRICS = [("dso", "DSO"), ("dio", "DIO"), ("dpo", "DPO"), ("ccc", "CCC")]

_GREEN = "63BE7B"
_YELLOW = "FFEB84"
_RED = "F8696B"


def export(heatmap: dict[str, list], filename: str, out_dir: Path | None = None) -> Path:
    """Write the heatmap matrix to an .xlsx with per-row colour scales.

    Args:
        heatmap: Output of ``working_capital.heatmap_data`` — parallel lists
            keyed by 'quarters', 'dso', 'dio', 'dpo', 'ccc'.
        filename: Output filename (e.g. "INFY_wc_heatmap.xlsx").
        out_dir: Output directory. Defaults to the configured Excel export dir.

    Returns:
        Path to the written workbook.

    Raises:
        ValueError: If the heatmap has no periods.
    """
    periods: list[str] = heatmap.get("quarters", [])
    if not periods:
        raise ValueError("Heatmap export needs at least one period")

    directory = out_dir or Path(_cfg["output_dir"])
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WC Heatmap"

    bold = Font(bold=True)
    ws.cell(1, 1, "Metric").font = bold
    for col, period in enumerate(periods, start=2):
        cell = ws.cell(1, col, period)
        cell.font = bold
        ws.column_dimensions[get_column_letter(col)].width = 12

    for row_offset, (key, label) in enumerate(_METRICS):
        row = 2 + row_offset
        ws.cell(row, 1, label).font = bold
        for col, value in enumerate(heatmap[key], start=2):
            ws.cell(row, col, round(float(value), 1))

        # Independent colour scale per metric row: low days green, high red.
        data_range = f"B{row}:{get_column_letter(1 + len(periods))}{row}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color=_GREEN,
                mid_type="percentile", mid_value=50, mid_color=_YELLOW,
                end_type="max", end_color=_RED,
            ),
        )

    wb.save(path)
    logger.info("WC heatmap exported: %s", path)
    return path
