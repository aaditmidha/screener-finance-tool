"""Template-style Excel model workbook (PL/BS/CF/Quarterly + Notes tabs).

Mirrors the hand-built reference models' canonical layout (see e.g.
"CG Power Consolidated Model.xlsx"): a derived P&L with Gross Profit, EBITDA,
EBIT, margins and YoY growth; raw BS/CF/Quarterly statements; and **Notes PL /
Notes BS / Notes CF** sheets grouping the granular child rows fetched from
Screener's expand API. Headers are bold and frozen, numbers use Indian-style
formats, percent rows render as percentages.

What it cannot do yet: match full annual-report depth (note-level schedules
like gratuity splits or MSME payables exist only in the AR PDFs — that is the
AR-pipeline roadmap). It exports everything Screener exposes, which is the
aggregated statements plus one level of expansion.
"""

import io
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from screener.config import CONFIG
from screener.models import ar_insights, operational
from screener.scraper.parser import CompanyFinancials, FinancialTable

logger = logging.getLogger(__name__)

_cfg = CONFIG["exporters"]["excel"]

_NUM_FMT = "#,##0;(#,##0)"
_PCT_FMT = "0.0%"

_TITLE_FONT = Font(bold=True, size=13)
_BOLD = Font(bold=True)


@dataclass
class _Row:
    """One output row: a label, per-period values, and styling hints."""

    label: str
    values: list[float | None]
    bold: bool = False
    pct: bool = False       # values are ratios → render as percentages
    is_header: bool = False  # group header — label only, no values
    num_fmt: str | None = None  # explicit Excel number format (overrides pct/num)


def _series(table: FinancialTable | None, *needles: str) -> list[float | None] | None:
    """Return the first matching row across *needles*, or None."""
    if table is None:
        return None
    for needle in needles:
        row = table.row(needle)
        if row:
            return row
    return None


def _combine(
    op: Callable[[float, float], float],
    a: list[float | None] | None,
    b: list[float | None] | None,
) -> list[float | None] | None:
    """Element-wise combine two series; None where either side is None."""
    if a is None or b is None:
        return None
    return [
        op(x, y) if x is not None and y is not None else None
        for x, y in zip(a, b)
    ]


def _ratio(num: list[float | None] | None, den: list[float | None] | None) -> list[float | None] | None:
    """Element-wise num/den with zero-denominator → None."""
    if num is None or den is None:
        return None
    return [
        (n / d) if n is not None and d not in (None, 0) else None
        for n, d in zip(num, den)
    ]


def _yoy(series: list[float | None] | None) -> list[float | None] | None:
    """Year-over-year growth of a series (first period None)."""
    if series is None:
        return None
    out: list[float | None] = [None]
    for prev, curr in zip(series, series[1:]):
        out.append((curr / prev - 1) if prev not in (None, 0) and curr is not None else None)
    return out


def _derived_pl_rows(fin: CompanyFinancials) -> list[_Row]:
    """Build the modelled P&L: revenue → margins → PAT, like the templates.

    Rows whose inputs Screener doesn't provide are simply omitted, so the
    sheet never shows fabricated zeros.
    """
    pl, notes = fin.profit_loss, fin.notes_pl
    if pl is None:
        return []
    n = len(pl.periods)

    revenue = _series(pl, "sales", "revenue")
    expenses = _series(pl, "expenses")
    op_profit = _series(pl, "operating profit")
    depreciation = _series(pl, "depreciation")
    interest = _series(pl, "interest")
    other_income = _series(pl, "other income")
    pbt = _series(pl, "profit before tax")
    tax_pct = _series(pl, "tax %")
    pat = _series(pl, "net profit")
    eps = _series(pl, "eps")

    # COGS from expand-API cost percentages when available.
    cogs = None
    if notes is not None and revenue is not None:
        material = _series(notes, "material cost")
        if material is not None:
            manufacturing = _series(notes, "manufacturing cost") or [0.0] * n
            cogs = [
                ((m or 0) + (f or 0)) / 100.0 * r if r is not None and m is not None else None
                for m, f, r in zip(material, manufacturing, revenue)
            ]
    employee = None
    other_costs = None
    if notes is not None and revenue is not None:
        emp_pct = _series(notes, "employee cost")
        oth_pct = _series(notes, "other cost")
        if emp_pct is not None:
            employee = [(p or 0) / 100.0 * r if r is not None and p is not None else None
                        for p, r in zip(emp_pct, revenue)]
        if oth_pct is not None:
            other_costs = [(p or 0) / 100.0 * r if r is not None and p is not None else None
                           for p, r in zip(oth_pct, revenue)]

    gross_profit = _combine(lambda r, c: r - c, revenue, cogs)
    ebit = _combine(lambda o, d: o - d, op_profit, depreciation)

    candidates: list[_Row | None] = [
        _Row("Revenue from Operations", revenue, bold=True) if revenue else None,
        _Row("YoY Growth %", _yoy(revenue), pct=True) if revenue else None,
        _Row("Total COGS (Material + Mfg, from cost %)", cogs) if cogs else None,
        _Row("Gross Profit", gross_profit, bold=True) if gross_profit else None,
        _Row("Gross Margin %", _ratio(gross_profit, revenue), pct=True) if gross_profit else None,
        _Row("Employee benefits expense (from cost %)", employee) if employee else None,
        _Row("Other expenses (from cost %)", other_costs) if other_costs else None,
        _Row("Total Expenses", expenses) if expenses else None,
        _Row("EBITDA (Operating Profit)", op_profit, bold=True) if op_profit else None,
        _Row("EBITDA Margin %", _ratio(op_profit, revenue), pct=True) if op_profit else None,
        _Row("Depreciation and amortisation", depreciation) if depreciation else None,
        _Row("EBIT", ebit, bold=True) if ebit else None,
        _Row("EBIT Margin %", _ratio(ebit, revenue), pct=True) if ebit else None,
        _Row("Finance costs", interest) if interest else None,
        _Row("Other income", other_income) if other_income else None,
        _Row("Profit Before Tax", pbt, bold=True) if pbt else None,
        _Row("Effective tax rate %", [t / 100.0 if t is not None else None for t in tax_pct],
             pct=True) if tax_pct else None,
        _Row("Profit After Tax", pat, bold=True) if pat else None,
        _Row("PAT Margin %", _ratio(pat, revenue), pct=True) if pat else None,
        _Row("EPS (₹)", eps) if eps else None,
    ]
    return [row for row in candidates if row is not None]


def _statement_rows(table: FinancialTable | None) -> list[_Row]:
    """Flat statement rows, percent-formatting any '%'-labelled lines."""
    if table is None:
        return []
    out: list[_Row] = []
    for label, values in table.rows.items():
        pct = "%" in label
        vals = [v / 100.0 if pct and v is not None else v for v in values]
        out.append(_Row(label, vals, pct=pct))
    return out


def _notes_rows(table: FinancialTable | None) -> list[_Row]:
    """Notes rows grouped under bold parent headers (labels 'Parent · Child')."""
    if table is None:
        return []
    out: list[_Row] = []
    current_parent = None
    for label, values in table.rows.items():
        parent, _, child = label.partition(" · ")
        if parent != current_parent:
            out.append(_Row(f"Note — {parent}", [], is_header=True))
            current_parent = parent
        pct = "%" in child
        vals = [v / 100.0 if pct and v is not None else v for v in values]
        out.append(_Row(child or label, vals, pct=pct))
    return out


def _write_sheet(ws: Worksheet, title: str, periods: list[str], rows: list[_Row]) -> None:
    """Write one formatted sheet: title, period header, styled data rows."""
    ws.cell(1, 1, title).font = _TITLE_FONT

    ws.cell(2, 1, "Line item").font = _BOLD
    for col, period in enumerate(periods, start=2):
        cell = ws.cell(2, col, period)
        cell.font = _BOLD
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["A"].width = 42
    ws.freeze_panes = "B3"

    r = 3
    for row in rows:
        label_cell = ws.cell(r, 1, row.label)
        if row.bold or row.is_header:
            label_cell.font = _BOLD
        if not row.is_header:
            for col, value in enumerate(row.values, start=2):
                if value is None:
                    continue
                cell = ws.cell(r, col, value)
                cell.number_format = row.num_fmt or (_PCT_FMT if row.pct else _NUM_FMT)
        r += 1


def _operational_rows(op: "object") -> list[_Row]:
    """Convert OperationalData metrics into formatted workbook rows.

    Args:
        op: A :class:`screener.models.operational.OperationalData`.

    Returns:
        One _Row per metric, with number formats matched to each unit
        (percent, turnover ``x``, or whole days).
    """
    fmt_map = {"pct": (_PCT_FMT, True), "x": ('0.00"x"', False), "days": ("0", False)}
    rows: list[_Row] = []
    for metric in op.metrics:
        num_fmt, pct = fmt_map.get(metric.fmt, (_NUM_FMT, False))
        rows.append(_Row(metric.label, list(metric.values), pct=pct, num_fmt=num_fmt))
    return rows


# Operational metrics surfaced on the Output Sheet (label → already in operational).
_OUTPUT_OP_METRICS = [
    "Asset turnover", "Receivable days", "Inventory days", "Payable days",
    "Cash conversion cycle (days)", "CFO / PAT",
]
# operational fmt → Excel number format / pct flag.
_OP_FMT = {"pct": (None, True), "x": ('0.00"x"', False), "days": ("0", False)}


def _output_rows(fin: CompanyFinancials) -> list[_Row]:
    """Build the Output Sheet — a one-look summary dashboard.

    Growth, margins, returns (ROCE/ROE), leverage and working-capital metrics
    across all periods, mirroring the reference models' Output/Summary sheet.

    Args:
        fin: Parsed company financials.

    Returns:
        Ordered _Row list; empty if there is no P&L.
    """
    pl, bs = fin.profit_loss, fin.balance_sheet
    if pl is None:
        return []
    rev = _series(pl, "sales", "revenue")
    op = _series(pl, "operating profit")
    dep = _series(pl, "depreciation")
    pat = _series(pl, "net profit", "profit after tax")
    eps = _series(pl, "eps")
    ebit = _combine(lambda o, d: o - d, op, dep)
    equity = _combine(lambda a, b: a + b,
                      _series(bs, "equity capital", "share capital"), _series(bs, "reserves"))
    debt = _series(bs, "borrowings", "debt")
    cap_employed = _combine(lambda e, d: e + d, equity, debt) if equity and debt else equity

    op_data = {m.label: (m.values, m.fmt) for m in operational.compute(fin).metrics}
    rows: list[_Row] = []

    def add(label, values, fmt="num", bold=False):
        if values and any(v is not None for v in values):
            num_fmt, pct = _OP_FMT.get(fmt, (None, fmt == "pct"))
            if fmt == "de":
                num_fmt = "0.00"
            rows.append(_Row(label, list(values), bold=bold, pct=pct, num_fmt=num_fmt))

    add("Revenue", rev, "num", bold=True)
    add("Revenue growth %", _yoy(rev), "pct")
    add("EBITDA", op, "num")
    add("EBITDA margin %", _ratio(op, rev), "pct")
    add("EBIT", ebit, "num")
    add("EBIT margin %", _ratio(ebit, rev), "pct")
    add("PAT", pat, "num", bold=True)
    add("PAT margin %", _ratio(pat, rev), "pct")
    add("EPS", eps, "num")
    add("ROCE %", _ratio(ebit, cap_employed), "pct")
    add("ROE %", _ratio(pat, equity), "pct")
    add("Debt / Equity", _ratio(debt, equity), "de")
    for label in _OUTPUT_OP_METRICS:
        if label in op_data:
            values, fmt = op_data[label]
            add(label, values, fmt)
    return rows


def _add_charts_sheet(wb: openpyxl.Workbook, company: str, fin: CompanyFinancials) -> None:
    """Add a Charts sheet with embedded focus-chart images, when plottable."""
    import io

    from openpyxl.drawing.image import Image as XLImage

    from screener.exporters import research_note

    periods, key_rows = research_note.key_financials(fin)
    pngs = research_note.focus_charts(periods, key_rows)
    if not pngs:
        return
    ws = wb.create_sheet(title="Charts")
    ws.cell(1, 1, f"{company} — Focus charts").font = _TITLE_FONT
    anchor = 3
    for title, png in pngs:
        ws.cell(anchor, 1, title).font = _BOLD
        image = XLImage(io.BytesIO(png))
        image.width, image.height = 480, 240
        ws.add_image(image, f"A{anchor + 1}")
        anchor += 15


def build_workbook(
    fin: CompanyFinancials,
    ar_rows: list | None = None,
    annual_rows: list | None = None,
) -> openpyxl.Workbook:
    """Assemble the full model workbook from parsed+enriched financials.

    Args:
        fin: Parsed company financials, ideally enriched with notes tables.
        ar_rows: Optional ARExtractedData rows → adds AR Financials / Screener-
            vs-AR / Risk Timeline sheets.
        annual_rows: Optional AnnualData rows → enables the discrepancy sheet.

    Returns:
        An openpyxl Workbook with PL/BS/CF/Quarterly/Operational/Notes sheets
        (plus AR sheets when ar_rows is given); sheets with no data are omitted.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    company = f"{fin.name} ({fin.symbol})" if fin.symbol else fin.name

    sheet_specs: list[tuple[str, list[str], list[_Row]]] = []
    if fin.profit_loss:
        sheet_specs.append(("Output Sheet", fin.profit_loss.periods, _output_rows(fin)))
        sheet_specs.append((
            "PL", fin.profit_loss.periods, _derived_pl_rows(fin)
        ))
    if fin.balance_sheet:
        sheet_specs.append(("BS", fin.balance_sheet.periods, _statement_rows(fin.balance_sheet)))
    if fin.cash_flow:
        sheet_specs.append(("CF", fin.cash_flow.periods, _statement_rows(fin.cash_flow)))
    if fin.quarters:
        sheet_specs.append(("Quarterly", fin.quarters.periods, _statement_rows(fin.quarters)))
    op = operational.compute(fin)
    if op.metrics:
        sheet_specs.append(("Operational Data", op.periods, _operational_rows(op)))
    if fin.notes_pl:
        sheet_specs.append(("Notes PL", fin.notes_pl.periods, _notes_rows(fin.notes_pl)))
    if fin.notes_bs:
        sheet_specs.append(("Notes BS", fin.notes_bs.periods, _notes_rows(fin.notes_bs)))
    if fin.notes_cf:
        sheet_specs.append(("Notes CF", fin.notes_cf.periods, _notes_rows(fin.notes_cf)))

    for name, periods, rows in sheet_specs:
        if not rows:
            continue
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, f"{company} — {name}", periods, rows)

    _add_charts_sheet(wb, company, fin)
    _append_ar_sheets(wb, company, ar_rows, annual_rows)

    if not wb.sheetnames:  # degenerate input — keep the workbook valid
        wb.create_sheet(title="Empty")
    logger.info("Model workbook built for %s: %s", fin.symbol, wb.sheetnames)
    return wb


# AR exact-figure rows: (display label, ARExtractedData attr).
_AR_SHEET_FIELDS = [
    ("Revenue", "revenue"), ("EBITDA", "ebitda"), ("PAT", "pat"), ("CFO", "cfo"),
    ("Capex", "capex"), ("Trade receivables", "trade_receivables"),
    ("Inventory", "inventory"), ("Trade payables", "trade_payables"),
    ("Total assets", "total_assets"), ("Total equity", "total_equity"),
    ("Total debt", "total_debt"), ("Cash", "cash"), ("Depreciation", "depreciation"),
    ("Interest expense", "interest_expense"), ("Tax expense", "tax_expense"),
]
_FILL_MODERATE = PatternFill("solid", fgColor="FFC000")
_FILL_LARGE = PatternFill("solid", fgColor="FF7C80")


def _write_table(ws: Worksheet, title: str, headers: list[str],
                 rows: list[list], fills: list | None = None) -> None:
    """Write a simple header + rows table, optionally filling whole rows."""
    ws.cell(1, 1, title).font = _TITLE_FONT
    for col, head in enumerate(headers, start=1):
        cell = ws.cell(2, col, head)
        cell.font = _BOLD
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(head) + 2)
    ws.freeze_panes = "A3"
    for r_off, row in enumerate(rows):
        fill = fills[r_off] if fills else None
        for col, value in enumerate(row, start=1):
            cell = ws.cell(3 + r_off, col, value)
            if fill is not None:
                cell.fill = fill


def _append_ar_sheets(wb: openpyxl.Workbook, company: str,
                      ar_rows: list | None, annual_rows: list | None) -> None:
    """Add the Annual-Report sheets when AR-extracted rows are available."""
    if not ar_rows:
        return
    ordered = sorted(ar_rows, key=lambda r: r.fiscal_year)
    years = [r.fiscal_year for r in ordered]

    # Sheet: exact AR figures (metric × year).
    rows = [
        _Row(label, [getattr(r, attr, None) for r in ordered])
        for label, attr in _AR_SHEET_FIELDS
        if any(getattr(r, attr, None) is not None for r in ordered)
    ]
    if rows:
        ws = wb.create_sheet(title="AR Financials")
        _write_sheet(ws, f"{company} — Annual Report (exact figures)",
                     [str(y) for y in years], rows)

    # Sheet: Screener vs AR discrepancy (colour-flagged).
    cells = ar_insights.discrepancies(ar_rows, annual_rows or [])
    if cells:
        table, fills = [], []
        for c in cells:
            table.append([c.metric, c.year, c.screener, c.ar,
                          round(c.diff_pct, 4) if c.diff_pct is not None else None, c.severity])
            fills.append(_FILL_LARGE if c.severity == "large"
                         else _FILL_MODERATE if c.severity == "moderate" else None)
        ws = wb.create_sheet(title="Screener vs AR")
        _write_table(ws, f"{company} — Screener vs Annual Report",
                     ["Metric", "Year", "Screener", "AR", "Diff %", "Severity"], table, fills)

    # Sheet: multi-year risk timeline.
    timeline = ar_insights.risk_timeline(ar_rows)
    if timeline:
        ws = wb.create_sheet(title="Risk Timeline")
        _write_table(ws, f"{company} — Key-risk timeline",
                     ["Risk", "First year", "Last year", "Frequency"],
                     [[e.risk, e.first_year, e.last_year, e.frequency] for e in timeline])


def to_bytes(fin: CompanyFinancials, ar_rows: list | None = None,
             annual_rows: list | None = None) -> bytes:
    """Return the model workbook as in-memory bytes (for download buttons).

    Args:
        fin: Parsed company financials.
        ar_rows: Optional ARExtractedData rows → adds AR sheets.
        annual_rows: Optional AnnualData rows → enables the discrepancy sheet.

    Returns:
        The .xlsx file contents.
    """
    buffer = io.BytesIO()
    build_workbook(fin, ar_rows=ar_rows, annual_rows=annual_rows).save(buffer)
    return buffer.getvalue()


def export(fin: CompanyFinancials, filename: str, out_dir: Path | None = None,
           ar_rows: list | None = None, annual_rows: list | None = None) -> Path:
    """Write the model workbook to disk and return its path.

    Args:
        fin: Parsed company financials.
        filename: Output filename.
        out_dir: Output directory; defaults to the configured Excel export dir.
        ar_rows: Optional ARExtractedData rows → adds AR sheets.
        annual_rows: Optional AnnualData rows → enables the discrepancy sheet.

    Returns:
        Path to the written workbook.
    """
    directory = out_dir or Path(_cfg["output_dir"])
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / filename
    build_workbook(fin, ar_rows=ar_rows, annual_rows=annual_rows).save(path)
    logger.info("Model workbook exported: %s", path)
    return path
