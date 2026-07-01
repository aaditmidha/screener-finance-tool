"""Tests for the template-style model workbook exporter."""

import io
import json
from dataclasses import dataclass

import openpyxl
import pytest

from screener.exporters import model_workbook
from screener.scraper import schedules
from screener.scraper.parser import parse_company_financials

_PAGE = """
<html><body><h1>CG Power</h1>
  <a href="/api/company/739/chart/"></a>
  <section id="profit-loss"><table>
    <thead><tr><th></th><th>Mar 2025</th><th>Mar 2026</th></tr></thead>
    <tbody>
      <tr><td>Sales</td><td>9,000</td><td>11,000</td></tr>
      <tr><td>Expenses +</td><td>7,800</td><td>9,500</td></tr>
      <tr><td>Operating Profit</td><td>1,200</td><td>1,500</td></tr>
      <tr><td>Depreciation</td><td>200</td><td>220</td></tr>
      <tr><td>Interest</td><td>50</td><td>40</td></tr>
      <tr><td>Profit before tax</td><td>1,050</td><td>1,360</td></tr>
      <tr><td>Tax %</td><td>25%</td><td>24%</td></tr>
      <tr><td>Net Profit</td><td>790</td><td>1,030</td></tr>
      <tr><td>EPS in Rs</td><td>5.2</td><td>6.7</td></tr>
    </tbody>
  </table></section>
  <section id="balance-sheet"><table>
    <thead><tr><th></th><th>Mar 2025</th><th>Mar 2026</th></tr></thead>
    <tbody>
      <tr><td>Equity Capital</td><td>200</td><td>200</td></tr>
      <tr><td>Reserves</td><td>3,800</td><td>4,800</td></tr>
      <tr><td>Borrowings +</td><td>900</td><td>700</td></tr>
      <tr><td>Other Assets +</td><td>3,100</td><td>4,200</td></tr>
      <tr><td>Total Assets</td><td>7,000</td><td>9,000</td></tr>
    </tbody>
  </table></section>
  <section id="cash-flow"><table>
    <thead><tr><th></th><th>Mar 2025</th><th>Mar 2026</th></tr></thead>
    <tbody>
      <tr><td>Cash from Operating Activity +</td><td>850</td><td>1,100</td></tr>
    </tbody>
  </table></section>
</body></html>
"""

_SCHEDULES = {
    "Expenses": {
        "Material Cost %": {"Mar 2025": "65%", "Mar 2026": "69%"},
        "Employee Cost %": {"Mar 2025": "6%", "Mar 2026": "8%"},
        "Other Cost %": {"Mar 2025": "8%", "Mar 2026": "9%"},
    },
    "Other Assets": {
        "Trade receivables": {"Mar 2025": "2,012", "Mar 2026": "2,924"},
        "Inventories": {"Mar 2025": "1,137", "Mar 2026": "1,584"},
    },
    "Cash from Operating Activity": {
        "Working capital changes": {"Mar 2025": "-300", "Mar 2026": "-498"},
    },
}


def _fake_fetch(url: str) -> str:
    for parent, payload in _SCHEDULES.items():
        if parent.replace(" ", "%20") in url:
            return json.dumps(payload)
    return json.dumps({})


@pytest.fixture()
def enriched_fin():
    fin = parse_company_financials(_PAGE)
    return schedules.enrich(fin, _PAGE, fetch_json=_fake_fetch)


@pytest.fixture()
def workbook(enriched_fin) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(model_workbook.to_bytes(enriched_fin)))


class TestSheetStructure:
    def test_canonical_sheets_present(self, workbook) -> None:
        """The download must mirror the reference template's sheet family."""
        for name in ("PL", "BS", "CF", "Notes PL", "Notes BS", "Notes CF"):
            assert name in workbook.sheetnames, f"missing sheet {name}"

    def test_periods_in_header_row(self, workbook) -> None:
        ws = workbook["PL"]
        headers = [ws.cell(2, c).value for c in (2, 3)]
        assert headers == ["Mar 2025", "Mar 2026"]

    def test_header_frozen(self, workbook) -> None:
        assert workbook["PL"].freeze_panes == "B3"


class TestDerivedPL:
    def _labels(self, ws) -> list[str]:
        return [ws.cell(r, 1).value for r in range(3, ws.max_row + 1)]

    def test_modelled_rows_present(self, workbook) -> None:
        labels = self._labels(workbook["PL"])
        for expected in ("Revenue from Operations", "YoY Growth %", "Gross Profit",
                         "EBITDA (Operating Profit)", "EBIT", "Profit After Tax",
                         "PAT Margin %"):
            assert any(expected in (l or "") for l in labels), f"missing {expected}"

    def test_gross_profit_derived_from_cost_pct(self, workbook) -> None:
        """GP = revenue − material% × revenue: 11000 − 0.69×11000 = 3410."""
        ws = workbook["PL"]
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == "Gross Profit":
                assert ws.cell(r, 3).value == pytest.approx(3410.0)
                return
        pytest.fail("Gross Profit row not found")

    def test_yoy_growth_value(self, workbook) -> None:
        """YoY FY26 = 11000/9000 − 1 ≈ 22.2%, stored as a ratio."""
        ws = workbook["PL"]
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == "YoY Growth %":
                assert ws.cell(r, 3).value == pytest.approx(0.2222, abs=1e-3)
                assert ws.cell(r, 3).number_format == "0.0%"
                return
        pytest.fail("YoY row not found")


class TestNotesSheets:
    def test_notes_bs_has_receivables_under_parent_header(self, workbook) -> None:
        ws = workbook["Notes BS"]
        labels = [ws.cell(r, 1).value for r in range(3, ws.max_row + 1)]
        assert "Note — Other Assets" in labels
        assert "Trade receivables" in labels

    def test_notes_values_written(self, workbook) -> None:
        ws = workbook["Notes BS"]
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == "Trade receivables":
                assert ws.cell(r, 3).value == pytest.approx(2924.0)
                return
        pytest.fail("Trade receivables row not found")


class TestOperationalSheet:
    def test_operational_sheet_present(self, workbook) -> None:
        assert "Operational Data" in workbook.sheetnames

    def test_operational_rows_and_formats(self, workbook) -> None:
        ws = workbook["Operational Data"]
        labels = [ws.cell(r, 1).value for r in range(3, ws.max_row + 1)]
        assert any("margin" in (l or "").lower() for l in labels)
        # A percent metric cell must carry a percent number format.
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == "EBITDA margin %":
                assert ws.cell(r, 3).number_format == "0.0%"
                return
        pytest.fail("EBITDA margin row not found")


class TestArSheets:
    @dataclass
    class _AR:
        fiscal_year: int
        revenue: float | None = None
        pat: float | None = None
        cfo: float | None = None
        total_assets: float | None = None
        total_debt: float | None = None
        total_equity: float | None = None
        ebitda: float | None = None
        capex: float | None = None
        trade_receivables: float | None = None
        inventory: float | None = None
        trade_payables: float | None = None
        cash: float | None = None
        depreciation: float | None = None
        interest_expense: float | None = None
        tax_expense: float | None = None
        guided_revenue_growth: float | None = None
        key_risks: object = None

    @dataclass
    class _Annual:
        fiscal_year_end: object
        revenue: float | None = None
        net_income: float | None = None
        total_assets: float | None = None
        total_debt: float | None = None
        shareholders_equity: float | None = None

    def _book(self, enriched_fin):
        import io
        import json
        from datetime import date
        ar_rows = [
            self._AR(2024, revenue=9000, pat=900, total_assets=12000,
                     trade_receivables=1500, key_risks=json.dumps(["Input cost inflation"])),
            self._AR(2025, revenue=11000, pat=600, total_assets=13000,
                     trade_receivables=1600, key_risks=json.dumps(["Input cost inflation", "Forex"])),
        ]
        annual_rows = [self._Annual(date(2025, 3, 31), revenue=11000, net_income=1200,
                                    total_assets=13000)]
        data = model_workbook.to_bytes(enriched_fin, ar_rows=ar_rows, annual_rows=annual_rows)
        return openpyxl.load_workbook(io.BytesIO(data))

    def test_ar_sheets_present(self, enriched_fin) -> None:
        wb = self._book(enriched_fin)
        for name in ("AR Financials", "Screener vs AR", "Risk Timeline"):
            assert name in wb.sheetnames

    def test_discrepancy_flagged(self, enriched_fin) -> None:
        wb = self._book(enriched_fin)
        ws = wb["Screener vs AR"]
        rows = [[ws.cell(r, c).value for c in range(1, 7)] for r in range(3, ws.max_row + 1)]
        pat = next(r for r in rows if r[0] == "Net profit / PAT")
        # AR PAT 600 vs Screener 1200 → -50% → large
        assert pat[5] == "large"

    def test_no_ar_rows_omits_sheets(self, enriched_fin) -> None:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(model_workbook.to_bytes(enriched_fin)))
        assert "AR Financials" not in wb.sheetnames


class TestOutputAndChartsSheets:
    def test_output_sheet_first_and_has_summary_rows(self, workbook) -> None:
        assert workbook.sheetnames[0] == "Output Sheet"
        ws = workbook["Output Sheet"]
        labels = [ws.cell(r, 1).value for r in range(3, ws.max_row + 1)]
        # Sectioned analyst summary: section titles + key derived rows.
        for expected in ("Income statement", "Revenue from operations",
                         "Common-size (% of revenue)", "EBITDA margin", "Growth (% YoY)",
                         "Balance sheet", "Ratios & returns", "ROCE", "ROE", "Debt / equity"):
            assert any(expected == l for l in labels), f"missing {expected}"

    def test_output_roce_value(self, workbook) -> None:
        ws = workbook["Output Sheet"]
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == "ROCE":
                # FY26: EBIT 1500-220=1280; capital employed 200+4800+700=5700 → 22.5%
                assert ws.cell(r, 3).value == pytest.approx(0.2246, abs=1e-3)
                assert ws.cell(r, 3).number_format == "0.0%"
                return
        pytest.fail("ROCE row not found")

    def test_charts_sheet_has_images(self, workbook) -> None:
        assert "Charts" in workbook.sheetnames
        assert len(workbook["Charts"]._images) >= 1   # embedded focus charts


class TestPeerSheet:
    def test_peer_sheet_added_from_dataframe(self, enriched_fin) -> None:
        import pandas as pd
        df = pd.DataFrame(
            {"name": ["Acme", "Beta"], "roce": [0.30, 0.20], "roe": [0.25, 0.18],
             "revenue_growth": [0.15, 0.10], "rank_composite": [1, 2]},
            index=["ACME", "BETA"])
        wb = openpyxl.load_workbook(io.BytesIO(model_workbook.to_bytes(enriched_fin, peer_df=df)))
        assert "Peer Comparison" in wb.sheetnames
        ws = wb["Peer Comparison"]
        symbols = [ws.cell(r, 1).value for r in range(3, ws.max_row + 1)]
        assert "ACME" in symbols and "BETA" in symbols
        # ROCE column is percent-formatted.
        assert ws.cell(3, 3).number_format == "0.0%"

    def test_no_peer_df_omits_sheet(self, enriched_fin) -> None:
        wb = openpyxl.load_workbook(io.BytesIO(model_workbook.to_bytes(enriched_fin)))
        assert "Peer Comparison" not in wb.sheetnames


class TestRobustness:
    def test_unenriched_fin_still_exports(self) -> None:
        """Without notes the workbook still has the statement sheets."""
        fin = parse_company_financials(_PAGE)
        wb = openpyxl.load_workbook(io.BytesIO(model_workbook.to_bytes(fin)))
        assert "PL" in wb.sheetnames
        assert "Notes PL" not in wb.sheetnames

    def test_empty_fin_yields_valid_workbook(self) -> None:
        fin = parse_company_financials("<html><body></body></html>")
        data = model_workbook.to_bytes(fin)
        assert data[:2] == b"PK"

    def test_export_writes_file(self, enriched_fin, tmp_path) -> None:
        path = model_workbook.export(enriched_fin, "cg.xlsx", out_dir=tmp_path)
        assert path.exists()
