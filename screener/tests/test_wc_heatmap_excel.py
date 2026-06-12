"""Tests for the Excel working-capital heatmap exporter."""

from pathlib import Path

import openpyxl
import pytest

from screener.exporters import wc_heatmap_excel

_HEATMAP = {
    "quarters": ["FY24 Q1", "FY24 Q2", "FY24 Q3"],
    "dso": [34.7, 34.3, 33.2],
    "dio": [106.2, 105.3, 100.9],
    "dpo": [64.5, 64.2, 64.5],
    "ccc": [76.4, 75.3, 69.6],
}


def test_export_writes_workbook(tmp_path: Path) -> None:
    path = wc_heatmap_excel.export(_HEATMAP, "hm.xlsx", out_dir=tmp_path)
    assert path.exists()
    assert path.read_bytes()[:2] == b"PK"


def test_matrix_layout_and_values(tmp_path: Path) -> None:
    path = wc_heatmap_excel.export(_HEATMAP, "hm.xlsx", out_dir=tmp_path)
    ws = openpyxl.load_workbook(path)["WC Heatmap"]

    assert [c.value for c in ws[1]] == ["Metric", "FY24 Q1", "FY24 Q2", "FY24 Q3"]
    assert [ws.cell(r, 1).value for r in range(2, 6)] == ["DSO", "DIO", "DPO", "CCC"]
    assert ws.cell(5, 4).value == pytest.approx(69.6)   # CCC, Q3


def test_colour_scale_applied_per_metric_row(tmp_path: Path) -> None:
    path = wc_heatmap_excel.export(_HEATMAP, "hm.xlsx", out_dir=tmp_path)
    ws = openpyxl.load_workbook(path)["WC Heatmap"]
    # One conditional-formatting colour scale per metric row.
    assert len(list(ws.conditional_formatting)) == 4


def test_empty_heatmap_raises(tmp_path: Path) -> None:
    with pytest.raises(ValueError):
        wc_heatmap_excel.export({"quarters": []}, "x.xlsx", out_dir=tmp_path)
