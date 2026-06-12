"""Tests for the Excel exporter."""

from pathlib import Path

import openpyxl
import pytest

from screener.exporters import excel


@pytest.fixture(autouse=True)
def _redirect_output(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Write exports into a temp dir instead of the configured one."""
    monkeypatch.setitem(excel._cfg, "output_dir", str(tmp_path))


def test_export_writes_xlsx(tmp_path: Path) -> None:
    path = excel.export({"Sheet1": [{"a": 1, "b": 2}]}, "t.xlsx")
    assert path.exists()
    assert path.read_bytes()[:2] == b"PK"   # xlsx is a zip archive


def test_sheet_headers_and_values_roundtrip(tmp_path: Path) -> None:
    rows = [{"metric": "revenue", "fy24": 153670.0},
            {"metric": "pat", "fy24": 26233.0}]
    path = excel.export({"PL": rows}, "pl.xlsx")

    wb = openpyxl.load_workbook(path)
    ws = wb["PL"]
    assert [c.value for c in ws[1]] == ["metric", "fy24"]
    assert ws.cell(2, 1).value == "revenue"
    assert ws.cell(3, 2).value == 26233.0


def test_long_sheet_name_truncated_to_31_chars(tmp_path: Path) -> None:
    name = "X" * 40
    path = excel.export({name: [{"a": 1}]}, "long.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["X" * 31]


def test_empty_rows_yield_blank_sheet(tmp_path: Path) -> None:
    path = excel.export({"Empty": []}, "empty.xlsx")
    wb = openpyxl.load_workbook(path)
    assert "Empty" in wb.sheetnames
    assert wb["Empty"].max_row == 1         # no data written
