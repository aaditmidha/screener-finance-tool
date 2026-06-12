"""Regenerate the CG Power model workbook live, end to end."""
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from screener.database.models import Base
from screener.exporters import model_workbook
from screener.scraper.acquisition import CompanyDataService

engine = create_engine("sqlite:///data/screener.db", future=True)
Base.metadata.create_all(engine)
with Session(engine) as session:
    service = CompanyDataService(session)
    fin = service.refresh("CGPOWER", force=True)

path = model_workbook.export(fin, "CGPOWER_model.xlsx")
print(f"written: {path}")

wb = openpyxl.load_workbook(path, read_only=True)
for ws in wb.worksheets:
    labels = [r[0].value for r in ws.iter_rows(min_col=1, max_col=1, max_row=60) if r[0].value]
    print(f"\n[{ws.title}] {ws.max_row}x{ws.max_column}")
    print(" | ".join(str(l)[:40] for l in labels[:25]))
wb.close()
